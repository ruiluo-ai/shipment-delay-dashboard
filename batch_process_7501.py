import os
from openpyxl import Workbook, load_workbook
from pdf_extract import extract_7501_data

pdf_folder = "pdfs"
excel_file = "7501_tracker_batch.xlsx"

if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "7501 Tracker"
    ws.append([
        "File Name",
        "Entry Number",
        "Summary Date",
        "Entry Date",
        "Port Code",
        "Status",
        "Error Message"
    ])

processed_files = set()

for row in ws.iter_rows(min_row=2,values_only=True):
    processed_files.add(row[0])

for file_name in os.listdir(pdf_folder):
    if file_name.lower().endswith(".pdf"):

        if file_name in processed_files:
            print(f"Skipping already processed file: {file_name}")
            continue
        pdf_path = os.path.join(pdf_folder,file_name)

        try:
            data = extract_7501_data(pdf_path)

            ws.append([
                    file_name,
                    data.get("Entry Number"),
                    data.get("Summary Date"),
                    data.get("Entry Date"),
                    data.get("Port Code"),
                    "Success",
                    ""
             ])
        except Exception as e:
            ws.append([
                file_name,
                None,
                None,
                None,
                None,
                "Failed",
                str(e)
            ])
wb.save(excel_file)

print("Batch processing complete.")