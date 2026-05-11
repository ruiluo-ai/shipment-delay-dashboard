from openpyxl import Workbook,load_workbook
from pdf_extract import extract_7501_data
import os
data = extract_7501_data("8SP-0066153-9-Corrected(Air-Shipment).pdf")
excel_file = "7501_tracker.xlsx"

if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "7501 Tracker"

    ws.append(["Entry Number","Summary Date","Entry Date","Port Code"])

ws.append([
    data["Entry Number"],
    data["Summary Date"],
    data["Entry Date"],
    data["Port Code"]
])

wb.save(excel_file)

print("Excel updated successfully.")